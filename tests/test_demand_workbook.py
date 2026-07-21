from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.demand_workbook import (
    DEMAND_COLUMNS,
    create_demo_optimization,
    example_demand_dataset,
    example_demand_dataframe,
    export_optimized_workbook,
    parse_demand_upload,
)


def test_demo_optimization_has_fixed_before_after_counts():
    result = create_demo_optimization(example_demand_dataset(), "DQN-GNN")

    assert result.before_conflict_count == 10
    assert result.after_conflict_count == 3
    assert len(result.suggestions) == 23
    assert sum(value.startswith("建议调整为 ") for value in result.suggestions.values()) == 7
    assert result.reduction_pct == 70.0


def test_parse_demand_upload_requires_the_suggestion_column(tmp_path):
    path = tmp_path / "bad.xlsx"
    example_demand_dataframe().drop(columns=["建议"]).to_excel(path, index=False)

    with pytest.raises(ValueError, match="缺少必填列.*建议"):
        parse_demand_upload(path.read_bytes(), path.name)


def test_export_preserves_uploaded_title_and_populates_suggestions(tmp_path):
    source = tmp_path / "需求表.xlsx"
    with pd.ExcelWriter(source) as writer:
        pd.DataFrame([["xx用频需求表（数据未验证）"]]).to_excel(
            writer, index=False, header=False
        )
        example_demand_dataframe().to_excel(writer, index=False, startrow=1)

    dataset = parse_demand_upload(source.read_bytes(), source.name)
    result = create_demo_optimization(dataset, "遗传算法")
    workbook = load_workbook(BytesIO(export_optimized_workbook(result)), data_only=True)
    sheet = workbook.active

    assert sheet.cell(1, 1).value == "xx用频需求表（数据未验证）"
    assert [sheet.cell(2, column).value for column in range(1, 13)] == list(DEMAND_COLUMNS)
    assert sheet.cell(3, 3).value == "短波通信系统"
    assert sheet.cell(3, 12).value.startswith(("建议调整为 ", "保持原工作频率"))
