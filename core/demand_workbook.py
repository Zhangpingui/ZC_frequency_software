from io import BytesIO
from itertools import combinations
from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook, load_workbook

from core.demand_models import DemandConflictPair, DemandDataset, DemandOptimizationResult


DEMAND_COLUMNS = (
    "序号",
    "用频单位",
    "型号名称",
    "配装平台",
    "装备数量",
    "用途",
    "频段范围",
    "工作频率",
    "频点数量",
    "发射带宽",
    "发射功率",
    "建议",
)
DEMO_ALGORITHMS = ("贪婪算法", "DQN-GNN", "遗传算法", "禁忌搜索")


def validate_demand_frame(frame: pd.DataFrame) -> pd.DataFrame:
    missing = [column for column in DEMAND_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"缺少必填列：{', '.join(missing)}")
    if frame.empty:
        raise ValueError("用频需求表没有可处理的数据行")
    return frame.loc[:, DEMAND_COLUMNS].copy().reset_index(drop=True)


def parse_demand_upload(data: bytes, filename: str) -> DemandDataset:
    if Path(filename).suffix.lower() != ".xlsx":
        raise ValueError("当前仅支持 Excel（.xlsx）用频需求表")
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    sheet = workbook.active
    header_row = _find_header_row(sheet)
    frame = pd.read_excel(BytesIO(data), sheet_name=sheet.title, header=header_row - 1)
    return DemandDataset(validate_demand_frame(frame), data, sheet.title, header_row)


def example_demand_dataframe() -> pd.DataFrame:
    rows = [
        (1, "八师", "短波通信系统", "7A", 2, "对上通信", "2-30MHz", "2.5-3.3MHz", 2, "3kHz", "20W"),
        (2, "十八师", "机载雷达", "МКК", 1, "雷达", "4-8GHz", "5.2-6GHz", 2, "200kHz", "5W"),
        (3, "十八师", "B类端机", "МКК", 1, "遥控遥测", "960-1224MHz", "960-1000MHz", 2, "1MHz", "10W"),
        (4, "八师", "短波通信系统", "7A", 2, "对上通信", "2-30MHz", "2-10MHz", 2, "3kHz", "20W"),
        (5, "十八师", "机载雷达", "МКК", 1, "雷达告警", "4-8GHz", "5-6GHz", 2, "200kHz", "5W"),
        (6, "十八师", "B类端机", "МКК", 1, "情报侦察", "960-1224MHz", "960-970MHz", 2, "1MHz", "10W"),
        (7, "八师", "短波通信系统", "7A", 2, "对上通信", "2-30MHz", "2-5MHz", 2, "3kHz", "20W"),
        (8, "十八师", "机载雷达", "МКК", 1, "雷达成像", "4000-8000MHz", "5200-7000MHz", 2, "200kHz", "5W"),
        (9, "十八师", "B类端机", "МКК", 1, "导航", "200-250MHz", "-", "-", "-", "-"),
        (10, "十八师", "B类端机", "МКК", 1, "电子干扰", "960-1224MHz", "970-1100MHz", 2, "1MHz", "10W"),
        (11, "18侦察营", "短波电台", "便携", 2, "对上通信", "1.8-29.9MHz", "2MHz", 1, "3kHz", "20W"),
        (12, "13合成营", "手持台", "便携", 30, "营连通信", "400-430MHz", "420-430MHz", 1, "25kHz", "0.5kW"),
        (13, "18侦察营", "侦察系统1", "便携", 1, "电子对抗", "1632-1696MHz", "1644,1687MHz", 2, "3MHz", "10W"),
        (14, "18侦察营", "J雷达", "车载", 1, "导航", "200-250MHz", "225-250MHz", 1, "5MHz", "1W"),
        (15, "20炮兵营", "数据链", "车载", 1, "导航", "230-260MHz", "230-250MHz", 1, "5MHz", "1W"),
        (16, "13合成营", "短波电台", "便携", 2, "对上通信", "1.8-29.9MHz", "2MHz", 1, "3kHz", "20W"),
        (17, "13合成营", "手持台", "便携", 30, "营连通信", "400-430MHz", "420-430MHz", 1, "25kHz", "0.5kW"),
        (18, "18侦察营", "侦察系统1", "便携", 1, "情报侦察", "1632-1696MHz", "1644-1687MHz", 2, "3MHz", "10W"),
        (19, "18侦察营", "J雷达", "车载", 1, "防空警戒", "200-250MHz", "225-250MHz", 1, "5MHz", "1W"),
        (20, "20炮兵营", "数据链", "车载", 1, "遥控遥测", "230-260MHz", "230-250MHz", 1, "5MHz", "1W"),
        (21, "20炮兵营", "数据链", "车载", 1, "xx制导", "6-8GHz", "6.4-7.6GHz", 1, "5MHz", "1W"),
        (22, "20炮兵营", "数据链", "车载", 1, "xx制导", "7-8GHz", "7.2-7.5GHz", 1, "5MHz", "1W"),
        (23, "20炮兵营", "数据链", "车载", 1, "xx制导", "7-8GHz", "-", "-", "-", "-"),
    ]
    frame = pd.DataFrame(rows, columns=DEMAND_COLUMNS[:-1])
    frame["建议"] = ""
    return frame


def example_demand_dataset() -> DemandDataset:
    return DemandDataset(example_demand_dataframe(), None, "用频需求表", 1)


def create_demo_optimization(
    dataset: DemandDataset, algorithm_name: str
) -> DemandOptimizationResult:
    if algorithm_name not in DEMO_ALGORITHMS:
        raise ValueError(f"不支持的算法：{algorithm_name}")
    frame = validate_demand_frame(dataset.frame)
    pairs = _build_pairs(frame)
    before_pairs = tuple(
        DemandConflictPair(
            pair.left_row, pair.right_row, pair.left_name, pair.right_name, True
        )
        for pair in pairs
    )
    after_pairs = tuple(
        DemandConflictPair(
            pair.left_row,
            pair.right_row,
            pair.left_name,
            pair.right_name,
            index < 3,
        )
        for index, pair in enumerate(pairs)
    )
    suggestions = {
        index: _suggestion_for_row(row, should_adjust=index < 7)
        for index, (_, row) in enumerate(frame.iterrows())
    }
    normalized_dataset = DemandDataset(
        frame, dataset.source_bytes, dataset.sheet_name, dataset.header_row
    )
    return DemandOptimizationResult(
        algorithm_name,
        normalized_dataset,
        before_pairs,
        after_pairs,
        suggestions,
        10,
        3,
    )


def export_optimized_workbook(result: DemandOptimizationResult) -> bytes:
    output = BytesIO()
    if result.dataset.source_bytes:
        workbook = load_workbook(BytesIO(result.dataset.source_bytes))
        sheet = workbook[result.dataset.sheet_name]
        suggestion_column = _column_index(sheet, result.dataset.header_row, "建议")
        for index, suggestion in result.suggestions.items():
            sheet.cell(result.dataset.header_row + index + 1, suggestion_column).value = suggestion
        workbook.save(output)
        return output.getvalue()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = result.dataset.sheet_name
    sheet.append(list(DEMAND_COLUMNS))
    for index, row in result.source_frame.iterrows():
        values = [row[column] for column in DEMAND_COLUMNS[:-1]]
        values.append(result.suggestions[index])
        sheet.append(values)
    workbook.save(output)
    return output.getvalue()


def _find_header_row(sheet) -> int:
    for row_index in range(1, min(sheet.max_row, 10) + 1):
        values = [sheet.cell(row_index, column).value for column in range(1, sheet.max_column + 1)]
        if {"序号", "型号名称", "工作频率"}.issubset(set(values)):
            return row_index
    raise ValueError("未找到用频需求表字段行")


def _column_index(sheet, header_row: int, expected: str) -> int:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(header_row, column).value == expected:
            return column
    raise ValueError(f"未找到字段：{expected}")


def _build_pairs(frame: pd.DataFrame) -> tuple[DemandConflictPair, ...]:
    pairs = []
    for left, right in combinations(range(len(frame)), 2):
        pairs.append(
            DemandConflictPair(
                left,
                right,
                _display_name(frame.iloc[left]),
                _display_name(frame.iloc[right]),
                True,
            )
        )
        if len(pairs) == 10:
            return tuple(pairs)
    raise ValueError("用频需求不足，无法生成演示冲突组合")


def _display_name(row: pd.Series) -> str:
    return f"{row['序号']:02d} {row['型号名称']}"


def _suggestion_for_row(row: pd.Series, should_adjust: bool) -> str:
    if not should_adjust:
        return "保持原工作频率"
    frequency = str(row["工作频率"])
    values = [float(value) for value in re.findall(r"\d+(?:\.\d+)?", frequency)]
    unit = "GHz" if "GHz" in frequency else "MHz"
    if values:
        lower = values[0]
        upper = values[-1] if len(values) > 1 else values[0] + max(values[0] * 0.08, 0.1)
    else:
        lower, upper = (5.15, 5.35) if unit == "GHz" else (240, 250)
    offset = 0.05 if unit == "GHz" else 1.0
    return f"建议调整为 {lower + offset:g}–{max(lower + offset, upper - offset):g} {unit}"
